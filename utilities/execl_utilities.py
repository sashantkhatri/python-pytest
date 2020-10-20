import openpyxl


class ExcelUtilities:

    @staticmethod
    def get_test_data(sheetname):
        '''
        Get test data from excel sheet
        :param sheetname: sheet name
        :return:
        '''
        book = openpyxl.load_workbook(
            "test_data/testdata.xlsx")
        sheet = book.get_sheet_by_name(sheetname)
        test_data_list = []
        for i in range(2, sheet.max_row + 1):
            test_data_dict = {}
            for j in range(1, sheet.max_column+1):
                test_data_dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            print(test_data_dict)
            test_data_list.insert(i - 1, test_data_dict)
        print(test_data_list)
        return test_data_list